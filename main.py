"""
飞书器材管理系统 - FastAPI 主应用
"""
from fastapi import FastAPI, Request, Response, Depends, HTTPException, status, Query, Body
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_, desc
from datetime import datetime, timedelta, timezone
from typing import Optional, List
from zoneinfo import ZoneInfo, available_timezones
import re
import jwt
import hashlib
import time
import secrets
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

import database
from database import (
    get_db, User, Category, EquipmentModel, Equipment, OperationLog
)
from config import (
    APP_ID, APP_SECRET, QR_CODE_PREFIX
)

# 兼容旧代码的变量名
FEISHU_APP_ID = APP_ID
FEISHU_JSAPI_URL = "https://lf-package-cn.feishucdn.com/obj/feishu-static/lark/op/h5-js-sdk-1.5.35.js"
SECRET_KEY = "your-secret-key-change-in-production"
SESSION_EXPIRE_DAYS = 7
DEFAULT_TIMEZONE = "Asia/Shanghai"
from feishu_auth import (
    get_user_info_by_code,
    get_or_create_user,
    get_user_by_id,
    get_jsapi_ticket
)


def get_request_timezone(request: Request) -> ZoneInfo:
    """从请求头读取用户时区。"""
    tz_name = (request.headers.get("X-Timezone") or DEFAULT_TIMEZONE).strip()
    try:
        return ZoneInfo(tz_name)
    except Exception:
        return ZoneInfo(DEFAULT_TIMEZONE)


def to_utc_naive(value: Optional[datetime], source_tz: ZoneInfo) -> Optional[datetime]:
    """将用户时区时间转换为 UTC 无时区 datetime（用于数据库存储/查询）。"""
    if value is None:
        return None
    if value.tzinfo is None:
        aware = value.replace(tzinfo=source_tz)
    else:
        aware = value.astimezone(source_tz)
    return aware.astimezone(timezone.utc).replace(tzinfo=None)


def from_utc_naive(value: Optional[datetime], target_tz: ZoneInfo) -> Optional[datetime]:
    """将数据库 UTC 无时区 datetime 转换为用户时区时间。"""
    if value is None:
        return None
    if value.tzinfo is None:
        aware_utc = value.replace(tzinfo=timezone.utc)
    else:
        aware_utc = value.astimezone(timezone.utc)
    return aware_utc.astimezone(target_tz)


def format_dt_in_timezone(value: Optional[datetime], target_tz: ZoneInfo, fmt: str) -> str:
    local_dt = from_utc_naive(value, target_tz)
    return local_dt.strftime(fmt) if local_dt else ""


CIRCULATION_ACTION_TYPES = ["CHECKOUT", "ASSIGN", "TRANSFER"]
ASSIGN_HOLDER_PATTERN = re.compile(r"\[分配给:([^\]]+)\]")
TRANSFER_HOLDER_PATTERN = re.compile(r"交接给([^，,]+)")


def extract_holder_name_from_purpose(
    purpose: Optional[str],
    action_type: Optional[str]
) -> Optional[str]:
    if not purpose:
        return None

    assign_match = ASSIGN_HOLDER_PATTERN.search(purpose)
    if assign_match:
        holder = assign_match.group(1).strip()
        if holder:
            return holder

    if action_type == "TRANSFER":
        transfer_match = TRANSFER_HOLDER_PATTERN.search(purpose)
        if transfer_match:
            holder = transfer_match.group(1).strip()
            if holder:
                return holder

    return None


def infer_holder_name_for_non_checkin_log(
    log: OperationLog,
    operator_name: Optional[str]
) -> Optional[str]:
    parsed_holder = extract_holder_name_from_purpose(log.purpose, log.action_type)
    if parsed_holder:
        return parsed_holder

    if log.action_type in ["CHECKOUT", "ASSIGN"]:
        return operator_name

    return None


async def get_user_name_by_open_id(
    feishu_open_id: Optional[str],
    db: AsyncSession,
    user_name_cache: dict[str, Optional[str]]
) -> Optional[str]:
    if not feishu_open_id:
        return None
    if feishu_open_id in user_name_cache:
        return user_name_cache[feishu_open_id]

    result = await db.execute(
        select(User.name).where(User.feishu_open_id == feishu_open_id)
    )
    user_name = result.scalar_one_or_none()
    user_name_cache[feishu_open_id] = user_name
    return user_name


async def resolve_log_holder_name(
    log: OperationLog,
    operator_name: Optional[str],
    db: AsyncSession,
    holder_cache: dict[int, str],
    user_name_cache: dict[str, Optional[str]]
) -> str:
    if log.id in holder_cache:
        return holder_cache[log.id]

    holder_name = infer_holder_name_for_non_checkin_log(log, operator_name)

    if not holder_name and log.action_type == "CHECKIN" and log.equipment_id:
        previous_log_result = await db.execute(
            select(OperationLog)
            .where(
                and_(
                    OperationLog.equipment_id == log.equipment_id,
                    OperationLog.action_type.in_(CIRCULATION_ACTION_TYPES),
                    OperationLog.id < log.id
                )
            )
            .order_by(desc(OperationLog.id))
            .limit(1)
        )
        previous_log = previous_log_result.scalar_one_or_none()
        if previous_log:
            previous_operator_name = await get_user_name_by_open_id(
                previous_log.feishu_open_id,
                db,
                user_name_cache
            )
            holder_name = await resolve_log_holder_name(
                previous_log,
                previous_operator_name,
                db,
                holder_cache,
                user_name_cache
            )

    resolved = holder_name or ""
    holder_cache[log.id] = resolved
    return resolved

# ============ Pydantic 模型 ============

class CheckoutAssignee(BaseModel):
    equipment_id: int
    user_id: int


class CheckoutRequest(BaseModel):
    equipment_ids: List[int]
    purpose: str
    expected_return_at: Optional[datetime] = None
    assignees: Optional[List[CheckoutAssignee]] = None

class CheckinRequest(BaseModel):
    equipment_ids: List[int]

class TransferRequest(BaseModel):
    equipment_ids: List[int]
    receiver_user_id: int
    transfer_reason: str
    expected_return_at: Optional[datetime] = None

class ScanRequest(BaseModel):
    qr_code: str
    mode: str = "checkout"


# ============ 分类管理 Pydantic 模型 ============

class CategoryCreate(BaseModel):
    name: str
    sort_order: int = 0

class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    sort_order: Optional[int] = None

class CategoryResponse(BaseModel):
    id: int
    name: str
    sort_order: int
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


# ============ 器材型号 Pydantic 模型 ============

class EquipmentModelCreate(BaseModel):
    category_id: int
    name: str
    description: Optional[str] = None
    total_count: int = 0

class EquipmentModelUpdate(BaseModel):
    category_id: Optional[int] = None
    name: Optional[str] = None
    description: Optional[str] = None
    total_count: Optional[int] = None

class EquipmentModelResponse(BaseModel):
    id: int
    category_id: int
    name: str
    description: Optional[str] = None
    total_count: int
    available_count: int
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


# ============ 器材实例 Pydantic 模型 ============

class EquipmentCreate(BaseModel):
    model_id: int
    count: int = 1  # 批量创建数量
    serial_prefix: Optional[str] = None  # 序列号前缀

class EquipmentUpdate(BaseModel):
    model_id: Optional[int] = None
    serial_number: Optional[str] = None
    qr_code: Optional[str] = None
    status: Optional[int] = None

class EquipmentResponse(BaseModel):
    id: int
    model_id: int
    model_name: Optional[str] = None
    serial_number: Optional[str] = None
    qr_code: Optional[str] = None
    status: int
    current_user_id: Optional[int] = None
    current_user_name: Optional[str] = None
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


# ============ 操作日志 Pydantic 模型 ============

class LogFilter(BaseModel):
    action_type: Optional[str] = None
    equipment_id: Optional[int] = None
    feishu_open_id: Optional[str] = None
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None

class OperationLogResponse(BaseModel):
    id: int
    equipment_id: int
    equipment_name: Optional[str] = None
    feishu_open_id: str
    user_name: Optional[str] = None
    action_type: str
    purpose: Optional[str] = None
    expected_return_at: Optional[datetime] = None
    actual_return_at: Optional[datetime] = None
    created_at: datetime

    class Config:
        from_attributes = True

# 创建应用
app = FastAPI(
    title="飞书器材管理系统",
    description="基于飞书免登录的器材出入库管理系统",
    version="1.0.0"
)

# 配置 CORS - 允许前端 8000 端口和其他域名访问
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:8000",  # 前端开发服务器
        "http://localhost:8001",  # 后端本身
        "http://127.0.0.1:8000",
        "http://127.0.0.1:8001",
        "http://wulianxx.com:33212", 
        "https://wulianxx.com:33212"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],  # 暴露 Content-Disposition 头给前端
)

# 挂载静态文件
app.mount("/static", StaticFiles(directory="static"), name="static")

# 模板
templates = Jinja2Templates(directory="templates")


# ============ 认证依赖项 ============

async def get_current_user(
    request: Request,
    db: AsyncSession = Depends(get_db)
) -> Optional[User]:
    """
    获取当前登录用户
    从Cookie中读取token并验证
    """
    token = request.cookies.get("session_token")
    if not token:
        return None
    
    try:
        # 解码JWT token
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        user_id = payload.get("user_id")
        
        if not user_id:
            return None
        
        # 检查token是否过期
        exp = payload.get("exp")
        if exp and datetime.utcnow().timestamp() > exp:
            return None
        
        # 获取用户信息
        user = await get_user_by_id(db, user_id)
        return user
        
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None


async def require_auth(
    request: Request,
    db: AsyncSession = Depends(get_db)
) -> User:
    """
    要求用户必须登录
    未登录时抛出401异常
    """
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="请先登录",
            headers={"WWW-Authenticate": "Bearer"}
        )
    return user


def create_session_token(user_id: int) -> str:
    """
    创建用户session token (JWT)
    """
    expire = datetime.utcnow() + timedelta(days=SESSION_EXPIRE_DAYS)
    payload = {
        "user_id": user_id,
        "exp": expire,
        "iat": datetime.utcnow()
    }
    return jwt.encode(payload, SECRET_KEY, algorithm="HS256")


# ============ 页面路由 ============

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    """
    首页/入口页面
    检测是否在飞书客户端内，引导用户登录
    """
    return templates.TemplateResponse("index.html", {
        "request": request,
        "app_id": FEISHU_APP_ID,
        "jsapi_url": FEISHU_JSAPI_URL
    })


@app.get("/devices", response_class=HTMLResponse)
async def devices_page(
    request: Request,
    user: User = Depends(require_auth)
):
    """设备管理页面"""
    return templates.TemplateResponse("devices.html", {
        "request": request,
        "user": user
    })


@app.get("/checkout", response_class=HTMLResponse)
async def checkout_page(
    request: Request,
    user: User = Depends(require_auth)
):
    """出入库页面"""
    return templates.TemplateResponse("checkout.html", {
        "request": request,
        "user": user
    })


@app.get("/logs", response_class=HTMLResponse)
async def logs_page(
    request: Request,
    user: User = Depends(require_auth)
):
    """操作日志页面"""
    return templates.TemplateResponse("logs.html", {
        "request": request,
        "user": user
    })


# ============ 认证API ============

@app.post("/api/auth/login")
async def login(
    request: Request,
    response: Response,
    db: AsyncSession = Depends(get_db)
):
    """
    飞书登录
    接收飞书code，完成登录流程，返回用户信息
    支持测试模式：code = "test_code_for_development" 时创建测试用户
    
    Request Body:
        {"code": "飞书临时授权码"}
    
    Response:
        {
            "success": true,
            "user": {
                "id": 1,
                "name": "用户名",
                "avatar": "头像URL",
                "open_id": "飞书open_id"
            }
        }
    """
    try:
        body = await request.json()
        code = body.get("code")
        
        if not code:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="缺少code参数"
            )
        
        # 测试模式：开发环境使用
        if code == "test_code_for_development":
            # 查找或创建测试用户
            result = await db.execute(
                select(User).where(User.feishu_open_id == "test_open_id")
            )
            user = result.scalar_one_or_none()
            
            if not user:
                user = User(
                    feishu_open_id="test_open_id",
                    feishu_user_id="test_user_id",
                    name="测试用户",
                    avatar_url=""
                )
                db.add(user)
                await db.commit()
                await db.refresh(user)
            
            # 创建session token
            token = create_session_token(user.id)
            
            # 设置Cookie - 开发环境使用 lax，生产环境使用 none + secure
            # localhost 不同端口共享 cookie，不需要 SameSite=None
            response.set_cookie(
                key="session_token",
                value=token,
                httponly=True,
                secure=False,
                samesite="lax",
                path="/",
                max_age=SESSION_EXPIRE_DAYS * 24 * 60 * 60
            )
            
            return {
                "success": True,
                "user": {
                    "id": user.id,
                    "name": user.name,
                    "avatar": user.avatar_url,
                    "open_id": user.feishu_open_id,
                    "user_id": user.feishu_user_id
                }
            }
        
        # 1. 用code换取飞书用户信息
        feishu_user = await get_user_info_by_code(code)
        
        if not feishu_user.get("open_id"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="获取飞书用户信息失败"
            )
        
        # 2. 创建或更新本地用户
        user = await get_or_create_user(
            db,
            feishu_open_id=feishu_user["open_id"],
            name=feishu_user.get("name"),
            avatar=feishu_user.get("avatar"),
            feishu_user_id=feishu_user.get("user_id")
        )
        
        # 3. 创建session token
        token = create_session_token(user.id)
        
        # 4. 设置Cookie - 开发环境使用 lax
        response.set_cookie(
            key="session_token",
            value=token,
            httponly=True,
            secure=False,  # 生产环境改为True
            samesite="lax",
            path="/",
            max_age=SESSION_EXPIRE_DAYS * 24 * 60 * 60
        )
        
        return {
            "success": True,
            "user": {
                "id": user.id,
                "name": user.name,
                "avatar": user.avatar_url,
                "open_id": user.feishu_open_id,
                "user_id": user.feishu_user_id
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"登录失败: {str(e)}"
        )


@app.get("/api/auth/config")
async def get_feishu_config():
    """
    获取飞书登录配置
    
    返回前端调用飞书JSAPI所需的配置信息
    
    Response:
        {
            "app_id": "cli_xxx",
            "jsapi_url": "https://..."
        }
    """
    return {
        "app_id": FEISHU_APP_ID,
        "jsapi_url": FEISHU_JSAPI_URL
    }


@app.get("/api/auth/me")
async def get_me(
    user: User = Depends(require_auth)
):
    """
    获取当前登录用户信息
    
    Response:
        {
            "id": 1,
            "name": "用户名",
            "avatar": "头像URL",
            "open_id": "飞书open_id",
            "user_id": "飞书user_id"
        }
    """
    return {
        "id": user.id,
        "name": user.name,
        "avatar": user.avatar_url,
        "open_id": user.feishu_open_id,
        "user_id": user.feishu_user_id
    }


@app.post("/api/auth/logout")
async def logout(response: Response):
    """
    退出登录
    清除session cookie
    """
    response.delete_cookie(
        key="session_token",
        path="/"
    )
    return {"success": True, "message": "已退出登录"}


# ============ 健康检查 ============

@app.get("/api/health")
async def health_check():
    """健康检查接口"""
    return {
        "status": "healthy",
        "timestamp": datetime.utcnow().isoformat()
    }


@app.get("/api/timezones")
async def get_timezones():
    """返回可用时区列表（IANA）"""
    zones = sorted(available_timezones())
    return {
        "total": len(zones),
        "timezones": zones
    }


# ============ 启动事件 ============

@app.on_event("startup")
async def startup_event():
    """应用启动时初始化数据库"""
    await database.init_db()
    print("✅ 数据库初始化完成")


# ============ 出入库API ============

@app.post("/api/checkout")
async def checkout_equipment(
    request: Request,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    批量出库
    
    Request Body:
        {
            "equipment_ids": [1, 2, 3],
            "purpose": "拍摄宣传片",
            "expected_return_at": "2025-02-20T18:00:00"
        }
    
    操作:
    1. 检查所有器材是否都在库
    2. 更新器材status=1(借出)
    3. 设置current_user_id
    4. 创建operation_logs记录
    5. 更新equipment_models的available_count
    """
    try:
        body = await request.json()
        checkout_req = CheckoutRequest(**body)
        user_tz = get_request_timezone(request)
        expected_return_at_utc = to_utc_naive(checkout_req.expected_return_at, user_tz)
        
        if not checkout_req.equipment_ids:
            raise HTTPException(status_code=400, detail="未选择器材")

        assignee_map: dict[int, int] = {}
        user_lookup: dict[int, User] = {user.id: user}
        requested_equipment_ids = set(checkout_req.equipment_ids)
        is_assign_operation = bool(checkout_req.assignees)
        now = datetime.utcnow()

        if checkout_req.assignees:
            for assignee in checkout_req.assignees:
                if assignee.equipment_id in requested_equipment_ids:
                    assignee_map[assignee.equipment_id] = assignee.user_id

            target_user_ids = set(assignee_map.values())
            if target_user_ids:
                users_result = await db.execute(
                    select(User).where(User.id.in_(list(target_user_ids)))
                )
                assignee_users = users_result.scalars().all()
                for assignee_user in assignee_users:
                    user_lookup[assignee_user.id] = assignee_user

                missing_user_ids = sorted(target_user_ids - set(user_lookup.keys()))
                if missing_user_ids:
                    raise HTTPException(
                        status_code=400,
                        detail=f"分配用户不存在: {', '.join(str(uid) for uid in missing_user_ids)}"
                    )
        
        results = {"success": [], "failed": []}
        
        for equipment_id in checkout_req.equipment_ids:
            # 查询器材
            result = await db.execute(
                select(Equipment, EquipmentModel)
                .join(EquipmentModel, Equipment.model_id == EquipmentModel.id)
                .where(Equipment.id == equipment_id)
            )
            row = result.first()
            
            if not row:
                results["failed"].append({
                    "id": equipment_id,
                    "reason": "器材不存在"
                })
                continue
            
            equipment, model = row
            
            if equipment.status == 1:
                results["failed"].append({
                    "id": equipment_id,
                    "name": model.name,
                    "reason": "该器材已借出"
                })
                continue
            
            # 更新器材状态
            receiver_user_id = assignee_map.get(equipment_id, user.id)
            receiver_user = user_lookup.get(receiver_user_id)

            equipment.status = 1
            equipment.current_user_id = receiver_user_id
            equipment.updated_at = now
            
            # 更新型号可用数量
            if model.available_count > 0:
                model.available_count -= 1
            
            # 创建操作日志
            log_purpose = checkout_req.purpose
            if receiver_user and receiver_user.id != user.id:
                log_purpose = f"{checkout_req.purpose} [分配给:{receiver_user.name}]"

            log = OperationLog(
                equipment_id=equipment_id,
                feishu_open_id=user.feishu_open_id,
                action_type="ASSIGN" if is_assign_operation else "CHECKOUT",
                purpose=log_purpose,
                expected_return_at=expected_return_at_utc,
                created_at=now
            )
            db.add(log)
            
            results["success"].append({
                "id": equipment_id,
                "name": model.name,
                "serial_number": equipment.serial_number,
                "assigned_user_id": receiver_user_id,
                "assigned_user_name": receiver_user.name if receiver_user else None
            })
        
        await db.commit()
        
        # 如果所有设备都失败，返回错误
        if len(results['success']) == 0:
            raise HTTPException(
                status_code=400, 
                detail=f"出库失败：{len(results['failed'])} 件器材无法出库，请检查设备状态"
            )
        
        return {
            "success": True,
            "data": results,
            "message": f"成功出库 {len(results['success'])} 件器材"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"出库失败: {str(e)}")


@app.post("/api/checkin")
async def checkin_equipment(
    request: Request,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    批量入库
    
    Request Body:
        {
            "equipment_ids": [1, 2, 3]
        }
    
    操作:
    1. 检查所有器材是否都已借出
    2. 更新器材status=0(在库)
    3. 清空current_user_id
    4. 更新operation_logs的actual_return_at
    5. 更新equipment_models的available_count
    """
    try:
        body = await request.json()
        checkin_req = CheckinRequest(**body)
        
        if not checkin_req.equipment_ids:
            raise HTTPException(status_code=400, detail="未选择器材")
        
        results = {"success": [], "failed": []}
        now = datetime.utcnow()
        
        for equipment_id in checkin_req.equipment_ids:
            # 查询器材
            result = await db.execute(
                select(Equipment, EquipmentModel)
                .join(EquipmentModel, Equipment.model_id == EquipmentModel.id)
                .where(Equipment.id == equipment_id)
            )
            row = result.first()
            
            if not row:
                results["failed"].append({
                    "id": equipment_id,
                    "reason": "器材不存在"
                })
                continue
            
            equipment, model = row
            
            if equipment.status == 0:
                results["failed"].append({
                    "id": equipment_id,
                    "name": model.name,
                    "reason": "该器材已在库"
                })
                continue
            
            # 更新器材状态
            equipment.status = 0
            equipment.current_user_id = None
            equipment.updated_at = now
            
            # 更新型号可用数量
            model.available_count += 1
            
            # 更新最近的出库记录的实际归还时间
            log_result = await db.execute(
                select(OperationLog)
                .where(
                    and_(
                        OperationLog.equipment_id == equipment_id,
                        OperationLog.action_type.in_(CIRCULATION_ACTION_TYPES),
                        OperationLog.actual_return_at.is_(None)
                    )
                )
                .order_by(desc(OperationLog.id))
                .limit(1)
            )
            last_log = log_result.scalar_one_or_none()
            
            if last_log:
                last_log.actual_return_at = now
            
            # 创建入库记录
            checkin_log = OperationLog(
                equipment_id=equipment_id,
                feishu_open_id=user.feishu_open_id,
                action_type="CHECKIN",
                actual_return_at=now,
                created_at=now
            )
            db.add(checkin_log)
            
            results["success"].append({
                "id": equipment_id,
                "name": model.name,
                "serial_number": equipment.serial_number
            })
        
        await db.commit()
        
        # 如果所有设备都失败，返回错误
        if len(results['success']) == 0:
            raise HTTPException(
                status_code=400, 
                detail=f"入库失败：{len(results['failed'])} 件器材无法入库，请检查设备状态"
            )
        
        return {
            "success": True,
            "data": results,
            "message": f"成功入库 {len(results['success'])} 件器材"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"入库失败: {str(e)}")


@app.post("/api/transfer")
async def transfer_equipment(
    request: Request,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    批量交接（借用人变更）
    """
    try:
        body = await request.json()
        transfer_req = TransferRequest(**body)
        user_tz = get_request_timezone(request)
        transfer_expected_return_at_utc = to_utc_naive(transfer_req.expected_return_at, user_tz)

        if not transfer_req.equipment_ids:
            raise HTTPException(status_code=400, detail="未选择器材")
        if not transfer_req.transfer_reason.strip():
            raise HTTPException(status_code=400, detail="交接原因不能为空")

        receiver_result = await db.execute(
            select(User).where(User.id == transfer_req.receiver_user_id)
        )
        receiver = receiver_result.scalar_one_or_none()
        if not receiver:
            raise HTTPException(status_code=404, detail="接收人不存在")

        results = {"success": [], "failed": []}
        now = datetime.utcnow()

        for equipment_id in transfer_req.equipment_ids:
            result = await db.execute(
                select(Equipment, EquipmentModel)
                .join(EquipmentModel, Equipment.model_id == EquipmentModel.id)
                .where(Equipment.id == equipment_id)
            )
            row = result.first()

            if not row:
                results["failed"].append({
                    "id": equipment_id,
                    "reason": "器材不存在"
                })
                continue

            equipment, model = row

            if equipment.status == 0:
                results["failed"].append({
                    "id": equipment_id,
                    "name": model.name,
                    "reason": "该器材当前在库，无法交接"
                })
                continue

            if not equipment.current_user_id:
                results["failed"].append({
                    "id": equipment_id,
                    "name": model.name,
                    "reason": "该器材当前无借用人，无法交接"
                })
                continue

            if equipment.current_user_id == transfer_req.receiver_user_id:
                results["failed"].append({
                    "id": equipment_id,
                    "name": model.name,
                    "reason": "接收人不能与原借用人相同"
                })
                continue

            previous_user_result = await db.execute(
                select(User).where(User.id == equipment.current_user_id)
            )
            previous_user = previous_user_result.scalar_one_or_none()
            previous_user_name = previous_user.name if previous_user else "未知用户"

            inherited_expected_return_at = None
            if transfer_req.expected_return_at is None:
                latest_non_null_result = await db.execute(
                    select(OperationLog)
                    .where(
                        and_(
                            OperationLog.equipment_id == equipment_id,
                            OperationLog.action_type.in_(CIRCULATION_ACTION_TYPES),
                            OperationLog.expected_return_at.is_not(None)
                        )
                    )
                    .order_by(desc(OperationLog.id))
                    .limit(1)
                )
                latest_non_null_log = latest_non_null_result.scalar_one_or_none()
                inherited_expected_return_at = (
                    latest_non_null_log.expected_return_at if latest_non_null_log else None
                )

            effective_expected_return_at = (
                transfer_expected_return_at_utc
                if transfer_req.expected_return_at is not None
                else inherited_expected_return_at
            )

            # 关闭上一条仍在进行中的借出/交接记录
            previous_active_log_result = await db.execute(
                select(OperationLog)
                .where(
                    and_(
                        OperationLog.equipment_id == equipment_id,
                        OperationLog.action_type.in_(CIRCULATION_ACTION_TYPES),
                        OperationLog.actual_return_at.is_(None)
                    )
                )
                .order_by(desc(OperationLog.id))
                .limit(1)
            )
            previous_active_log = previous_active_log_result.scalar_one_or_none()
            if previous_active_log:
                previous_active_log.actual_return_at = now

            equipment.current_user_id = transfer_req.receiver_user_id
            equipment.updated_at = now

            log = OperationLog(
                equipment_id=equipment_id,
                feishu_open_id=user.feishu_open_id,
                action_type="TRANSFER",
                purpose=f"从{previous_user_name}交接给{receiver.name}，原因：{transfer_req.transfer_reason.strip()}",
                expected_return_at=effective_expected_return_at,
                created_at=now
            )
            db.add(log)

            results["success"].append({
                "id": equipment_id,
                "name": model.name,
                "serial_number": equipment.serial_number
            })

        await db.commit()

        if len(results["success"]) == 0:
            raise HTTPException(
                status_code=400,
                detail=f"交接失败：{len(results['failed'])} 件器材无法交接，请检查设备状态"
            )

        return {
            "success": True,
            "data": results,
            "message": f"成功交接 {len(results['success'])} 件器材"
        }
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"交接失败: {str(e)}")


@app.get("/api/checkout/active")
async def get_active_checkouts(
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    获取当前借出列表
    
    Query参数:
        - user_id: 可选，筛选特定用户的借出记录
    
    返回所有status=1的器材列表，包含借用信息
    """
    try:
        # 查询所有借出状态的器材
        result = await db.execute(
            select(Equipment, EquipmentModel, User)
            .join(EquipmentModel, Equipment.model_id == EquipmentModel.id)
            .outerjoin(User, Equipment.current_user_id == User.id)
            .where(Equipment.status == 1)
        )
        rows = result.all()
        
        data = []
        for equipment, model, borrower in rows:
            # 获取最新的出库记录
            log_result = await db.execute(
                select(OperationLog)
                .where(
                    and_(
                        OperationLog.equipment_id == equipment.id,
                        OperationLog.action_type.in_(CIRCULATION_ACTION_TYPES)
                    )
                )
                .order_by(desc(OperationLog.id))
            )
            last_log = log_result.scalar_one_or_none()
            
            data.append({
                "equipment_id": equipment.id,
                "model_name": model.name if model else "",
                "serial_number": equipment.serial_number,
                "qr_code": equipment.qr_code,
                "current_user_id": equipment.current_user_id,
                "current_user_name": borrower.name if borrower else None,
                "checkout_time": last_log.created_at if last_log else None,
                "purpose": last_log.purpose if last_log else None,
                "expected_return_at": last_log.expected_return_at if last_log else None
            })
        
        return {
            "total": len(data),
            "data": data
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取借出列表失败: {str(e)}")


# ============ 飞书扫码配置API ============

@app.get("/api/feishu/scan-config")
async def get_scan_config(
    request: Request,
    page_url: Optional[str] = Query(None, description="当前网页URL，用于JSAPI签名")
):
    """
    返回飞书扫码JSAPI配置

    返回前端调用tt.scanCode所需的配置信息
    注意：签名有效期10分钟，前端不应缓存此配置
    """
    try:
        # 生成签名参数（使用毫秒级时间戳）
        timestamp = str(int(time.time() * 1000))  # 毫秒级时间戳
        nonce_str = secrets.token_hex(8)

        # JSAPI签名必须使用前端页面URL（不含hash），不能使用本接口URL
        if page_url:
            url = page_url.split("#")[0]
        else:
            # 兼容旧调用：未传page_url时退化为当前请求URL
            url = str(request.url).split("#")[0]

        # 获取JSAPI ticket（每次都重新获取）
        ticket = await get_jsapi_ticket()

        # 生成签名: sha1(jsapi_ticket + noncestr + timestamp + url)
        # 注意：参数名必须按字典序排序
        data = f"jsapi_ticket={ticket}&noncestr={nonce_str}&timestamp={timestamp}&url={url}"
        signature = hashlib.sha1(data.encode()).hexdigest()

        return {
            "app_id": APP_ID,
            "timestamp": timestamp,
            "nonceStr": nonce_str,  # 驼峰写法！
            "signature": signature,
            "scan_type": ["qrCode", "barCode"]
        }

    except Exception as e:
        # 如果获取ticket失败，返回错误而不是空签名
        raise HTTPException(
            status_code=500,
            detail=f"获取扫码配置失败: {str(e)}"
        )


# ============ 扫码处理API ============

@app.post("/api/scan/process")
async def process_scan(
    request: Request,
    db: AsyncSession = Depends(get_db)
):
    """
    处理扫码结果
    
    Request Body:
        {
            "qr_code": "扫码结果",
            "mode": "checkout" 或 "checkin"
        }
    
    解析qr_code，查询器材信息，检查状态
    支持多种匹配方式：条形码内容、设备ID、设备编号
    """
    try:
        body = await request.json()
        qr_code = body.get("qr_code", "").strip()
        mode = body.get("mode", "checkout")
        
        if not qr_code:
            return {
                "success": False,
                "message": "条形码内容不能为空",
                "can_proceed": False
            }
        
        row = None
        
        # 方式1：直接匹配条形码内容
        result = await db.execute(
            select(Equipment, EquipmentModel, User, Category)
            .join(EquipmentModel, Equipment.model_id == EquipmentModel.id)
            .outerjoin(Category, EquipmentModel.category_id == Category.id)
            .outerjoin(User, Equipment.current_user_id == User.id)
            .where(Equipment.qr_code == qr_code)
        )
        row = result.first()
        
        # 方式2：尝试解析条形码格式 EQUIP:ID:HASH
        if not row and qr_code.startswith(QR_CODE_PREFIX):
            try:
                parts = qr_code.split(":")
                if len(parts) >= 2:
                    equipment_id = int(parts[1])
                    result = await db.execute(
                        select(Equipment, EquipmentModel, User, Category)
                        .join(EquipmentModel, Equipment.model_id == EquipmentModel.id)
                        .outerjoin(Category, EquipmentModel.category_id == Category.id)
                        .outerjoin(User, Equipment.current_user_id == User.id)
                        .where(Equipment.id == equipment_id)
                    )
                    row = result.first()
            except (ValueError, IndexError):
                pass
        
        # 方式3：尝试按设备编号匹配（支持扫描条形码）
        if not row:
            result = await db.execute(
                select(Equipment, EquipmentModel, User, Category)
                .join(EquipmentModel, Equipment.model_id == EquipmentModel.id)
                .outerjoin(Category, EquipmentModel.category_id == Category.id)
                .outerjoin(User, Equipment.current_user_id == User.id)
                .where(Equipment.serial_number == qr_code)
            )
            row = result.first()
        
        # 方式4：尝试将扫码内容作为设备ID直接查询
        if not row:
            try:
                equipment_id = int(qr_code)
                result = await db.execute(
                    select(Equipment, EquipmentModel, User, Category)
                    .join(EquipmentModel, Equipment.model_id == EquipmentModel.id)
                    .outerjoin(Category, EquipmentModel.category_id == Category.id)
                    .outerjoin(User, Equipment.current_user_id == User.id)
                    .where(Equipment.id == equipment_id)
                )
                row = result.first()
            except (ValueError, IndexError):
                pass
        
        if not row:
            return {
                "success": False,
                "message": f"设备 '{qr_code}' 未登记，请先前往设备管理登记该设备",
                "can_proceed": False,
                "debug_info": {
                    "searched_qr_code": qr_code
                }
            }
        
        equipment, model, borrower, category = row
        
        response_data = {
            "success": True,
            "equipment": {
                "id": equipment.id,
                "model_name": model.name if model else "",
                "serial_number": equipment.serial_number,
                "qr_code": equipment.qr_code,
                "status": equipment.status,
                "current_user_id": equipment.current_user_id,
                "current_user_name": borrower.name if borrower else None,
                "category_name": category.name if category else "未分类"
            }
        }
        
        # 根据模式检查状态
        if mode == "checkout":
            # 出库模式：检查是否在库
            if equipment.status == 1:
                # 获取借出信息
                log_result = await db.execute(
                    select(OperationLog)
                    .where(
                        and_(
                            OperationLog.equipment_id == equipment.id,
                            OperationLog.action_type.in_(CIRCULATION_ACTION_TYPES)
                        )
                    )
                    .order_by(desc(OperationLog.id))
                    .limit(1)
                )
                last_log = log_result.scalar_one_or_none()
                
                response_data["can_proceed"] = False
                response_data["message"] = f"该器材已被 {borrower.name if borrower else '他人'} 借出"
                response_data["equipment"]["checkout_time"] = last_log.created_at if last_log else None
                response_data["equipment"]["purpose"] = last_log.purpose if last_log else None
            else:
                response_data["can_proceed"] = True
                response_data["message"] = "扫描成功，可以出库"
                
        elif mode in ["checkin", "transfer"]:
            # 入库模式：检查是否已借出
            if equipment.status == 0:
                response_data["can_proceed"] = False
                response_data["message"] = "该器材已在库，无法操作"
            else:
                # 获取借出信息
                log_result = await db.execute(
                    select(OperationLog)
                    .where(
                        and_(
                            OperationLog.equipment_id == equipment.id,
                            OperationLog.action_type.in_(CIRCULATION_ACTION_TYPES)
                        )
                    )
                    .order_by(desc(OperationLog.id))
                    .limit(1)
                )
                last_log = log_result.scalar_one_or_none()
                
                response_data["can_proceed"] = True
                response_data["message"] = "扫描成功，可以入库" if mode == "checkin" else "扫描成功，可以交接"
                response_data["equipment"]["checkout_time"] = last_log.created_at if last_log else None
                response_data["equipment"]["purpose"] = last_log.purpose if last_log else None
        
        return response_data
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"处理扫码失败: {str(e)}")


# ============ 器材查询API ============

@app.get("/api/equipments/by-qrcode")
async def get_equipment_by_qr(
    qr_code: str,
    db: AsyncSession = Depends(get_db)
):
    """通过条形码查询器材信息
    
    支持多种匹配方式：
    1. 直接匹配条形码内容
    2. 解析 EQUIP:ID:HASH 格式
    3. 匹配设备编号（serial_number）
    4. 直接作为设备ID查询
    """
    try:
        qr_code = qr_code.strip()
        row = None
        
        # 方式1：尝试直接匹配
        result = await db.execute(
            select(Equipment, EquipmentModel, User)
            .join(EquipmentModel, Equipment.model_id == EquipmentModel.id)
            .outerjoin(User, Equipment.current_user_id == User.id)
            .where(Equipment.qr_code == qr_code)
        )
        row = result.first()
        
        # 方式2：尝试解析条形码格式
        if not row and qr_code.startswith(QR_CODE_PREFIX):
            try:
                parts = qr_code.split(":")
                if len(parts) >= 2:
                    equipment_id = int(parts[1])
                    result = await db.execute(
                        select(Equipment, EquipmentModel, User)
                        .join(EquipmentModel, Equipment.model_id == EquipmentModel.id)
                        .outerjoin(User, Equipment.current_user_id == User.id)
                        .where(Equipment.id == equipment_id)
                    )
                    row = result.first()
            except (ValueError, IndexError):
                pass
        
        # 方式3：尝试按设备编号匹配
        if not row:
            result = await db.execute(
                select(Equipment, EquipmentModel, User)
                .join(EquipmentModel, Equipment.model_id == EquipmentModel.id)
                .outerjoin(User, Equipment.current_user_id == User.id)
                .where(Equipment.serial_number == qr_code)
            )
            row = result.first()
        
        # 方式4：尝试将内容作为设备ID查询
        if not row:
            try:
                equipment_id = int(qr_code)
                result = await db.execute(
                    select(Equipment, EquipmentModel, User)
                    .join(EquipmentModel, Equipment.model_id == EquipmentModel.id)
                    .outerjoin(User, Equipment.current_user_id == User.id)
                    .where(Equipment.id == equipment_id)
                )
                row = result.first()
            except (ValueError, IndexError):
                pass
        
        if not row:
            raise HTTPException(status_code=404, detail="器材不存在")
        
        equipment, model, borrower = row
        
        # 获取最近的出库记录
        checkout_time = None
        purpose = None
        if equipment.status == 1 and equipment.current_user_id:
            log_result = await db.execute(
                select(OperationLog)
                .where(
                    and_(
                        OperationLog.equipment_id == equipment.id,
                        OperationLog.action_type.in_(CIRCULATION_ACTION_TYPES)
                    )
                )
                .order_by(desc(OperationLog.id))
                .limit(1)
            )
            last_log = log_result.scalar_one_or_none()
            if last_log:
                checkout_time = last_log.created_at
                purpose = last_log.purpose
        
        return {
            "equipment_id": equipment.id,
            "model_name": model.name if model else "",
            "serial_number": equipment.serial_number,
            "qr_code": equipment.qr_code,
            "status": equipment.status,
            "current_user_id": equipment.current_user_id,
            "current_user_name": borrower.name if borrower else None,
            "checkout_time": checkout_time,
            "purpose": purpose
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"查询失败: {str(e)}")


# ============ 分类管理 API ============

@app.get("/api/categories", response_model=List[CategoryResponse])
async def get_categories(
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    获取所有分类
    按 sort_order 排序
    """
    try:
        result = await db.execute(
            select(Category).order_by(Category.sort_order, Category.id)
        )
        categories = result.scalars().all()
        return categories
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取分类失败: {str(e)}")


@app.post("/api/categories", response_model=CategoryResponse)
async def create_category(
    category_data: CategoryCreate,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    创建分类
    """
    try:
        # 检查分类名称是否已存在
        result = await db.execute(
            select(Category).where(Category.name == category_data.name)
        )
        if result.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="分类名称已存在")
        
        # 创建新分类
        category = Category(
            name=category_data.name,
            sort_order=category_data.sort_order
        )
        db.add(category)
        await db.commit()
        await db.refresh(category)
        
        # 创建操作日志
        log = OperationLog(
            equipment_id=None,
            feishu_open_id=user.feishu_open_id,
            action_type="CREATE_CATEGORY",
            purpose=f"创建分类: {category.name}"
        )
        db.add(log)
        await db.commit()
        
        return category
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"创建分类失败: {str(e)}")


@app.put("/api/categories/{category_id}", response_model=CategoryResponse)
async def update_category(
    category_id: int,
    category_data: CategoryUpdate,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    更新分类
    """
    try:
        # 获取分类
        result = await db.execute(
            select(Category).where(Category.id == category_id)
        )
        category = result.scalar_one_or_none()
        
        if not category:
            raise HTTPException(status_code=404, detail="分类不存在")
        
        # 记录变更内容
        changes = []
        
        # 检查新名称是否与其他分类冲突
        if category_data.name and category_data.name != category.name:
            result = await db.execute(
                select(Category).where(
                    and_(Category.name == category_data.name, Category.id != category_id)
                )
            )
            if result.scalar_one_or_none():
                raise HTTPException(status_code=400, detail="分类名称已存在")
            changes.append(f"名称: {category.name} → {category_data.name}")
            category.name = category_data.name
        
        if category_data.sort_order is not None and category_data.sort_order != category.sort_order:
            changes.append(f"排序: {category.sort_order} → {category_data.sort_order}")
            category.sort_order = category_data.sort_order
        
        await db.commit()
        await db.refresh(category)
        
        # 创建操作日志
        if changes:
            log = OperationLog(
                equipment_id=None,
                feishu_open_id=user.feishu_open_id,
                action_type="UPDATE_CATEGORY",
                purpose=f"编辑分类: {', '.join(changes)}"
            )
            db.add(log)
            await db.commit()
        
        return category
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"更新分类失败: {str(e)}")


@app.delete("/api/categories/{category_id}")
async def delete_category(
    category_id: int,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    删除分类
    会级联删除关联的型号和器材
    """
    try:
        # 获取分类
        result = await db.execute(
            select(Category).where(Category.id == category_id)
        )
        category = result.scalar_one_or_none()
        
        if not category:
            raise HTTPException(status_code=404, detail="分类不存在")
        
        category_name = category.name
        await db.delete(category)
        await db.commit()
        
        # 创建操作日志
        log = OperationLog(
            equipment_id=None,
            feishu_open_id=user.feishu_open_id,
            action_type="DELETE_CATEGORY",
            purpose=f"删除分类: {category_name}"
        )
        db.add(log)
        await db.commit()
        
        return {"success": True, "message": "分类已删除"}
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"删除分类失败: {str(e)}")


# ============ 分类排序 API ============

class CategoryReorderRequest(BaseModel):
    category_ids: List[int]

@app.post("/api/categories/reorder")
async def reorder_categories(
    request: CategoryReorderRequest,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    重新排序分类
    传入按新顺序排列的分类 ID 列表
    """
    try:
        for index, category_id in enumerate(request.category_ids):
            result = await db.execute(
                select(Category).where(Category.id == category_id)
            )
            category = result.scalar_one_or_none()
            if category:
                category.sort_order = index
        
        await db.commit()
        return {"success": True, "message": "分类排序已更新"}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"排序失败: {str(e)}")


# ============ 器材型号 API ============

@app.get("/api/models", response_model=List[EquipmentModelResponse])
async def get_models(
    category_id: Optional[int] = None,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    获取型号列表
    支持按 category_id 筛选
    """
    try:
        query = select(EquipmentModel)
        if category_id:
            query = query.where(EquipmentModel.category_id == category_id)
        query = query.order_by(EquipmentModel.sort_order, EquipmentModel.id)
        
        result = await db.execute(query)
        models = result.scalars().all()
        return models
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取型号列表失败: {str(e)}")


@app.get("/api/models/{model_id}", response_model=EquipmentModelResponse)
async def get_model_detail(
    model_id: int,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    获取型号详情
    """
    try:
        result = await db.execute(
            select(EquipmentModel).where(EquipmentModel.id == model_id)
        )
        model = result.scalar_one_or_none()
        
        if not model:
            raise HTTPException(status_code=404, detail="型号不存在")
        
        return model
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取型号详情失败: {str(e)}")


@app.post("/api/models", response_model=EquipmentModelResponse)
async def create_model(
    model_data: EquipmentModelCreate,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    创建器材型号
    """
    try:
        # 检查分类是否存在
        result = await db.execute(
            select(Category).where(Category.id == model_data.category_id)
        )
        if not result.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="所属分类不存在")
        
        # 创建型号
        model = EquipmentModel(
            category_id=model_data.category_id,
            name=model_data.name,
            description=model_data.description,
            total_count=model_data.total_count,
            available_count=0  # 初始为0，添加器材实例时更新
        )
        db.add(model)
        await db.commit()
        await db.refresh(model)
        
        # 创建操作日志
        log = OperationLog(
            equipment_id=None,
            feishu_open_id=user.feishu_open_id,
            action_type="CREATE_MODEL",
            purpose=f"添加型号: {model.name}"
        )
        db.add(log)
        await db.commit()
        
        return model
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"创建型号失败: {str(e)}")


@app.put("/api/models/{model_id}", response_model=EquipmentModelResponse)
async def update_model(
    model_id: int,
    model_data: EquipmentModelUpdate,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    更新器材型号
    """
    try:
        result = await db.execute(
            select(EquipmentModel).where(EquipmentModel.id == model_id)
        )
        model = result.scalar_one_or_none()
        
        if not model:
            raise HTTPException(status_code=404, detail="型号不存在")
        
        # 检查新分类是否存在
        if model_data.category_id:
            cat_result = await db.execute(
                select(Category).where(Category.id == model_data.category_id)
            )
            if not cat_result.scalar_one_or_none():
                raise HTTPException(status_code=400, detail="所属分类不存在")
            model.category_id = model_data.category_id
        
        # 记录变更内容
        changes = []
        
        if model_data.name is not None and model_data.name != model.name:
            changes.append(f"名称: {model.name} → {model_data.name}")
            model.name = model_data.name
        if model_data.description is not None and model_data.description != model.description:
            changes.append(f"描述变更")
            model.description = model_data.description
        if model_data.total_count is not None and model_data.total_count != model.total_count:
            changes.append(f"总数量: {model.total_count} → {model_data.total_count}")
            model.total_count = model_data.total_count
        
        await db.commit()
        await db.refresh(model)
        
        # 创建操作日志
        if changes:
            log = OperationLog(
                equipment_id=None,
                feishu_open_id=user.feishu_open_id,
                action_type="UPDATE_MODEL",
                purpose=f"编辑型号: {', '.join(changes)}"
            )
            db.add(log)
            await db.commit()
        
        return model
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"更新型号失败: {str(e)}")


@app.delete("/api/models/{model_id}")
async def delete_model(
    model_id: int,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    删除器材型号
    会级联删除关联的器材实例
    """
    try:
        result = await db.execute(
            select(EquipmentModel).where(EquipmentModel.id == model_id)
        )
        model = result.scalar_one_or_none()
        
        if not model:
            raise HTTPException(status_code=404, detail="型号不存在")
        
        model_name = model.name
        await db.delete(model)
        await db.commit()
        
        # 创建操作日志
        log = OperationLog(
            equipment_id=None,
            feishu_open_id=user.feishu_open_id,
            action_type="DELETE_MODEL",
            purpose=f"删除型号: {model_name}"
        )
        db.add(log)
        await db.commit()
        
        return {"success": True, "message": "型号已删除"}
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"删除型号失败: {str(e)}")


# ============ 型号排序 API ============

class ModelReorderRequest(BaseModel):
    category_id: int
    model_ids: List[int]

@app.post("/api/models/reorder")
async def reorder_models(
    request: ModelReorderRequest,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    重新排序型号（在指定分类内）
    传入按新顺序排列的型号 ID 列表
    """
    try:
        for index, model_id in enumerate(request.model_ids):
            result = await db.execute(
                select(EquipmentModel).where(
                    and_(EquipmentModel.id == model_id, EquipmentModel.category_id == request.category_id)
                )
            )
            model = result.scalar_one_or_none()
            if model:
                model.sort_order = index
        
        await db.commit()
        return {"success": True, "message": "型号排序已更新"}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"排序失败: {str(e)}")


# ============ 器材实例 API ============

@app.get("/api/equipments")
async def get_equipments(
    model_id: Optional[int] = None,
    status: Optional[int] = None,
    current_user_id: Optional[int] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    获取器材列表
    支持按 model_id, status, current_user_id 筛选
    支持分页
    """
    try:
        # 构建查询
        query = select(Equipment, EquipmentModel, User).join(
            EquipmentModel, Equipment.model_id == EquipmentModel.id
        ).outerjoin(
            User, Equipment.current_user_id == User.id
        )
        
        # 应用筛选条件
        if model_id:
            query = query.where(Equipment.model_id == model_id)
        if status is not None:
            query = query.where(Equipment.status == status)
        if current_user_id:
            query = query.where(Equipment.current_user_id == current_user_id)
        
        # 获取总数
        count_query = select(Equipment)
        if model_id:
            count_query = count_query.where(Equipment.model_id == model_id)
        if status is not None:
            count_query = count_query.where(Equipment.status == status)
        if current_user_id:
            count_query = count_query.where(Equipment.current_user_id == current_user_id)
        
        count_result = await db.execute(count_query)
        total = len(count_result.scalars().all())
        
        # 分页
        query = query.order_by(Equipment.sort_order, Equipment.id).offset((page - 1) * page_size).limit(page_size)
        result = await db.execute(query)
        rows = result.all()
        
        # 组装响应数据
        data = []
        for equipment, model, borrower in rows:
            # 获取出库详情（如果已借出）
            checkout_time = None
            purpose = None
            expected_return_at = None
            if equipment.status == 1 and equipment.current_user_id:
                log_result = await db.execute(
                    select(OperationLog)
                    .where(
                        and_(
                            OperationLog.equipment_id == equipment.id,
                            OperationLog.action_type.in_(CIRCULATION_ACTION_TYPES)
                        )
                    )
                    .order_by(desc(OperationLog.id))
                    .limit(1)
                )
                last_log = log_result.scalar_one_or_none()
                if last_log:
                    checkout_time = last_log.created_at.isoformat() if last_log.created_at else None
                    purpose = last_log.purpose
                    expected_return_at = last_log.expected_return_at.isoformat() if last_log.expected_return_at else None
            
            data.append({
                "id": equipment.id,
                "model_id": equipment.model_id,
                "model_name": model.name if model else None,
                "serial_number": equipment.serial_number,
                "qr_code": equipment.qr_code,
                "status": equipment.status,
                "current_user_id": equipment.current_user_id,
                "current_user_name": borrower.name if borrower else None,
                "checkout_time": checkout_time,
                "purpose": purpose,
                "expected_return_at": expected_return_at,
                "created_at": equipment.created_at,
                "updated_at": equipment.updated_at
            })
        
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "data": data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取器材列表失败: {str(e)}")


# ============ 设备导出 API ============

@app.get("/api/equipments/export")
async def export_equipments(
    request: Request,
    model_id: Optional[int] = None,
    status: Optional[int] = None,
    current_user_id: Optional[int] = None,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    导出设备列表为 Excel 文件
    支持筛选条件，导出所有符合条件的记录（不分页）
    """
    try:
        user_tz = get_request_timezone(request)
        # 构建查询 - 获取所有符合条件的记录
        query = select(Equipment, EquipmentModel, User).join(
            EquipmentModel, Equipment.model_id == EquipmentModel.id
        ).outerjoin(
            User, Equipment.current_user_id == User.id
        )
        
        # 应用筛选条件
        if model_id:
            query = query.where(Equipment.model_id == model_id)
        if status is not None:
            query = query.where(Equipment.status == status)
        if current_user_id:
            query = query.where(Equipment.current_user_id == current_user_id)
        
        # 排序并获取所有数据
        query = query.order_by(EquipmentModel.name, Equipment.serial_number)
        result = await db.execute(query)
        rows = result.all()
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "设备列表"
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        headers = ['分类', '型号', '设备编码', '设备唯一码', '状态', '借用人', '用途', '预计归还', '出库时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据
        for row_idx, (equipment, model, borrower) in enumerate(rows, 2):
            # 获取出库信息（如果已借出）
            checkout_time = None
            purpose = None
            expected_return_at = None
            if equipment.status == 1:
                log_result = await db.execute(
                    select(OperationLog).where(
                        and_(
                            OperationLog.equipment_id == equipment.id,
                            OperationLog.action_type.in_(CIRCULATION_ACTION_TYPES)
                        )
                    ).order_by(desc(OperationLog.id)).limit(1)
                )
                last_log = log_result.scalar_one_or_none()
                if last_log:
                    checkout_time = last_log.created_at
                    purpose = last_log.purpose
                    expected_return_at = last_log.expected_return_at
            
            # 状态映射
            status_map = {0: "在库", 1: "借出"}
            status_text = status_map.get(equipment.status, "未知")
            
            # 获取分类信息
            category_name = ""
            if model:
                cat_result = await db.execute(
                    select(Category).where(Category.id == model.category_id)
                )
                category = cat_result.scalar_one_or_none()
                category_name = category.name if category else ""
            
            data = [
                category_name,
                model.name if model else "",
                equipment.serial_number or "",
                equipment.qr_code or "",
                status_text,
                borrower.name if borrower else "",
                purpose or "",
                format_dt_in_timezone(expected_return_at, user_tz, '%Y-%m-%d'),
                format_dt_in_timezone(checkout_time, user_tz, '%Y-%m-%d %H:%M'),
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
        
        # 调整列宽
        column_widths = [15, 20, 20, 25, 10, 15, 30, 15, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        date_str = format_dt_in_timezone(datetime.utcnow(), user_tz, '%Y%m%d')
        filename = f"设备清单_{date_str}.xlsx"
        
        # 使用 urllib.parse.quote 处理中文文件名
        from urllib.parse import quote
        encoded_filename = quote(filename)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出设备失败: {str(e)}")


@app.get("/api/equipments/{equipment_id}")
async def get_equipment_detail(
    equipment_id: int,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    获取器材详情
    """
    try:
        result = await db.execute(
            select(Equipment, EquipmentModel, User).join(
                EquipmentModel, Equipment.model_id == EquipmentModel.id
            ).outerjoin(
                User, Equipment.current_user_id == User.id
            ).where(Equipment.id == equipment_id)
        )
        row = result.first()
        
        if not row:
            raise HTTPException(status_code=404, detail="器材不存在")
        
        equipment, model, borrower = row
        
        # 获取最近的出库记录
        checkout_time = None
        purpose = None
        expected_return_at = None
        if equipment.status == 1:
            log_result = await db.execute(
                select(OperationLog).where(
                    and_(
                        OperationLog.equipment_id == equipment.id,
                        OperationLog.action_type.in_(CIRCULATION_ACTION_TYPES)
                    )
                ).order_by(desc(OperationLog.id)).limit(1)
            )
            last_log = log_result.scalar_one_or_none()
            if last_log:
                checkout_time = last_log.created_at
                purpose = last_log.purpose
                expected_return_at = last_log.expected_return_at
        
        return {
            "id": equipment.id,
            "model_id": equipment.model_id,
            "model_name": model.name if model else None,
            "serial_number": equipment.serial_number,
            "qr_code": equipment.qr_code,
            "status": equipment.status,
            "current_user_id": equipment.current_user_id,
            "current_user_name": borrower.name if borrower else None,
            "checkout_time": checkout_time,
            "purpose": purpose,
            "expected_return_at": expected_return_at,
            "created_at": equipment.created_at,
            "updated_at": equipment.updated_at
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取器材详情失败: {str(e)}")


@app.post("/api/equipments")
async def create_equipments(
    request: Request,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    批量创建器材
    支持自定义设备编号，不自动添加后缀
    """
    try:
        body = await request.json()
        model_id = body.get("model_id")
        serial_numbers = body.get("serial_numbers", [])  # 设备编号列表
        
        if not model_id:
            raise HTTPException(status_code=400, detail="缺少型号ID")
        
        # 检查型号是否存在
        result = await db.execute(
            select(EquipmentModel).where(EquipmentModel.id == model_id)
        )
        model = result.scalar_one_or_none()
        
        if not model:
            raise HTTPException(status_code=400, detail="器材型号不存在")
        
        if not serial_numbers or len(serial_numbers) == 0:
            raise HTTPException(status_code=400, detail="设备编号不能为空")
        
        created_equipments = []
        
        for serial_number in serial_numbers:
            # 直接使用传入的设备编号，不做任何修改
            equipment = Equipment(
                model_id=model_id,
                serial_number=serial_number.strip() if serial_number else None,
                qr_code=None,  # 先创建获取ID后再更新
                status=0,
                current_user_id=None
            )
            db.add(equipment)
            await db.flush()  # 获取ID
            
            # 生成条形码
            qr_content = f"{QR_CODE_PREFIX}:{equipment.id}:{hashlib.md5(str(equipment.id).encode()).hexdigest()[:8]}"
            equipment.qr_code = qr_content
            
            created_equipments.append({
                "id": equipment.id,
                "serial_number": serial_number,
                "qr_code": qr_content
            })
            
            # 创建操作日志
            log = OperationLog(
                equipment_id=equipment.id,
                feishu_open_id=user.feishu_open_id,
                action_type="CREATE"
            )
            db.add(log)
        
        # 更新型号的总数量和可用数量
        model.total_count += len(serial_numbers)
        model.available_count += len(serial_numbers)
        
        await db.commit()
        
        return {
            "success": True,
            "count": len(created_equipments),
            "data": created_equipments
        }
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"创建器材失败: {str(e)}")


@app.put("/api/equipments/{equipment_id}")
async def update_equipment(
    equipment_id: int,
    equipment_data: EquipmentUpdate,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    更新器材信息
    """
    try:
        result = await db.execute(
            select(Equipment, EquipmentModel).join(
                EquipmentModel, Equipment.model_id == EquipmentModel.id
            ).where(Equipment.id == equipment_id)
        )
        row = result.first()
        
        if not row:
            raise HTTPException(status_code=404, detail="器材不存在")
        
        equipment, model = row
        old_status = equipment.status
        
        # 如果更换型号，更新型号的数量
        if equipment_data.model_id and equipment_data.model_id != equipment.model_id:
            # 检查新型号是否存在
            new_model_result = await db.execute(
                select(EquipmentModel).where(EquipmentModel.id == equipment_data.model_id)
            )
            new_model = new_model_result.scalar_one_or_none()
            if not new_model:
                raise HTTPException(status_code=400, detail="目标型号不存在")
            
            # 更新原型号数量
            model.total_count -= 1
            if equipment.status == 0:
                model.available_count -= 1
            
            # 更新新型号数量
            new_model.total_count += 1
            if equipment.status == 0:
                new_model.available_count += 1
            
            equipment.model_id = equipment_data.model_id
        
        # 记录变更内容
        changes = []
        
        if equipment_data.serial_number is not None and equipment_data.serial_number != equipment.serial_number:
            changes.append(f"编号: {equipment.serial_number or '无'} → {equipment_data.serial_number}")
            equipment.serial_number = equipment_data.serial_number
        if equipment_data.qr_code is not None and equipment_data.qr_code != equipment.qr_code:
            changes.append(f"条形码变更")
            equipment.qr_code = equipment_data.qr_code
        
        # 状态变更需要记录日志
        status_changed = False
        if equipment_data.status is not None and equipment_data.status != old_status:
            equipment.status = equipment_data.status
            status_changed = True
            
            # 创建状态变更日志
            action_type = "CHECKOUT" if equipment_data.status == 1 else "CHECKIN"
            log = OperationLog(
                equipment_id=equipment_id,
                feishu_open_id=user.feishu_open_id,
                action_type=action_type
            )
            db.add(log)
        
        # 如果有其他变更，记录UPDATE日志
        if changes and not status_changed:
            log = OperationLog(
                equipment_id=equipment_id,
                feishu_open_id=user.feishu_open_id,
                action_type="UPDATE",
                purpose=f"编辑设备: {', '.join(changes)}"
            )
            db.add(log)
        
        await db.commit()
        await db.refresh(equipment)
        
        return {
            "success": True,
            "message": "器材已更新",
            "equipment": {
                "id": equipment.id,
                "model_id": equipment.model_id,
                "serial_number": equipment.serial_number,
                "qr_code": equipment.qr_code,
                "status": equipment.status
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"更新器材失败: {str(e)}")


@app.delete("/api/equipments/{equipment_id}")
async def delete_equipment(
    equipment_id: int,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    删除器材
    如果器材已借出，不能删除
    """
    try:
        result = await db.execute(
            select(Equipment, EquipmentModel).join(
                EquipmentModel, Equipment.model_id == EquipmentModel.id
            ).where(Equipment.id == equipment_id)
        )
        row = result.first()
        
        if not row:
            raise HTTPException(status_code=404, detail="器材不存在")
        
        equipment, model = row
        
        if equipment.status == 1:
            raise HTTPException(status_code=400, detail="该器材当前已借出，无法删除")
        
        # 更新型号数量
        model.total_count -= 1
        model.available_count -= 1
        
        # 创建删除日志
        log = OperationLog(
            equipment_id=equipment_id,
            feishu_open_id=user.feishu_open_id,
            action_type="DELETE"
        )
        db.add(log)
        
        await db.delete(equipment)
        await db.commit()
        
        return {"success": True, "message": "器材已删除"}
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"删除器材失败: {str(e)}")


# ============ 设备排序 API ============

class DeviceReorderRequest(BaseModel):
    model_id: int
    device_ids: List[int]

@app.post("/api/equipments/reorder")
async def reorder_devices(
    request: DeviceReorderRequest,
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    重新排序设备（在指定型号内）
    传入按新顺序排列的设备 ID 列表
    """
    try:
        for index, device_id in enumerate(request.device_ids):
            result = await db.execute(
                select(Equipment).where(
                    and_(Equipment.id == device_id, Equipment.model_id == request.model_id)
                )
            )
            device = result.scalar_one_or_none()
            if device:
                device.sort_order = index
        
        await db.commit()
        return {"success": True, "message": "设备排序已更新"}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"排序失败: {str(e)}")


# ============ 日志导出 API ============

@app.get("/api/logs/export")
async def export_logs(
    request: Request,
    action_type: Optional[str] = None,
    equipment_id: Optional[int] = None,
    feishu_open_id: Optional[str] = None,
    serial_number: Optional[str] = None,
    model_name: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    operator_name: Optional[str] = None,
    holder_name: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    导出操作日志为 Excel 文件
    支持筛选条件，导出所有符合条件的记录（不分页）
    """
    try:
        user_tz = get_request_timezone(request)
        start_date_utc = to_utc_naive(start_date, user_tz)
        end_date_utc = to_utc_naive(end_date, user_tz)

        # 构建查询 - 获取所有符合条件的记录
        query = select(OperationLog, Equipment, EquipmentModel, User).outerjoin(
            Equipment, OperationLog.equipment_id == Equipment.id
        ).outerjoin(
            EquipmentModel, Equipment.model_id == EquipmentModel.id
        ).join(
            User, OperationLog.feishu_open_id == User.feishu_open_id
        )
        
        # 应用筛选条件
        # 所有编辑类操作类型（用于"编辑信息"筛选）
        edit_action_types = [
            "CREATE", "UPDATE", "DELETE",  # 设备操作
            "CREATE_CATEGORY", "UPDATE_CATEGORY", "DELETE_CATEGORY",  # 分类操作
            "CREATE_MODEL", "UPDATE_MODEL", "DELETE_MODEL"  # 型号操作
        ]
        checkout_action_types = ["CHECKOUT", "ASSIGN", "CHECKIN", "TRANSFER"]
        
        if action_type:
            # 处理组合筛选
            if action_type == "EDIT":
                query = query.where(OperationLog.action_type.in_(edit_action_types))
            elif action_type == "出入库":
                query = query.where(OperationLog.action_type.in_(checkout_action_types))
            elif action_type in ["交接", "TRANSFER"]:
                query = query.where(OperationLog.action_type == "TRANSFER")
            elif action_type in ["分配", "ASSIGN"]:
                query = query.where(OperationLog.action_type == "ASSIGN")
            else:
                # 单个操作类型筛选
                query = query.where(OperationLog.action_type == action_type)
        if equipment_id:
            query = query.where(OperationLog.equipment_id == equipment_id)
        if feishu_open_id:
            query = query.where(OperationLog.feishu_open_id == feishu_open_id)
        if serial_number:
            query = query.where(Equipment.serial_number.ilike(f"%{serial_number}%"))
        if model_name:
            query = query.where(EquipmentModel.name.ilike(f"%{model_name}%"))
        if start_date_utc:
            query = query.where(OperationLog.created_at >= start_date_utc)
        if end_date_utc:
            query = query.where(OperationLog.created_at <= end_date_utc)
        if operator_name:
            query = query.where(User.name.ilike(f"%{operator_name}%"))
        if search:
            # 搜索关键词匹配设备型号名称、设备编号或用途
            search_pattern = f"%{search}%"
            query = query.where(
                or_(
                    EquipmentModel.name.ilike(search_pattern),
                    Equipment.serial_number.ilike(search_pattern),
                    OperationLog.purpose.ilike(search_pattern)
                )
            )
        
        # 排序并获取所有数据
        query = query.order_by(desc(OperationLog.id))
        result = await db.execute(query)
        rows = result.all()
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "操作日志"
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        headers = ['时间', '操作人', '设备持有人', '操作类型', '设备信息', '入手时间', '用途', '预计归还日期', '实际出手时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 操作类型映射 - Excel导出用
        def map_action_type(action_type: str) -> str:
            mapping = {
                "CHECKOUT": "出库",
                "ASSIGN": "分配",
                "CHECKIN": "入库",
                "TRANSFER": "交接",
                "CREATE": "添加设备",
                "UPDATE": "修改设备",
                "DELETE": "删除设备",
                "CREATE_CATEGORY": "添加分类",
                "UPDATE_CATEGORY": "修改分类",
                "DELETE_CATEGORY": "删除分类",
                "CREATE_MODEL": "添加型号",
                "UPDATE_MODEL": "修改型号",
                "DELETE_MODEL": "删除型号",
            }
            return mapping.get(action_type, action_type)
        
        # 辅助函数：获取出库日期
        async def get_checkout_date(log, db):
            """获取出库日期：出库操作返回自身创建时间，入库操作查找对应的出库记录"""
            if log.action_type in CIRCULATION_ACTION_TYPES:
                return log.created_at
            elif log.action_type == "CHECKIN":
                # 查找该设备的上一次出库记录（在入库之前最近的出库）
                checkout_query = select(OperationLog).where(
                    OperationLog.equipment_id == log.equipment_id,
                    OperationLog.action_type.in_(CIRCULATION_ACTION_TYPES),
                    OperationLog.id < log.id
                ).order_by(desc(OperationLog.id)).limit(1)
                checkout_result = await db.execute(checkout_query)
                checkout_log = checkout_result.scalar_one_or_none()
                return checkout_log.created_at if checkout_log else None
            return None
        
        # 预计算持有人（并按持有人筛选）
        holder_filter_keyword = holder_name.strip().lower() if holder_name else None
        holder_name_cache: dict[int, str] = {}
        user_name_cache: dict[str, Optional[str]] = {}
        resolved_rows: list[tuple[OperationLog, Optional[Equipment], Optional[EquipmentModel], Optional[User], str]] = []
        for log, equipment, model, operator in rows:
            if operator and operator.feishu_open_id:
                user_name_cache[operator.feishu_open_id] = operator.name

            resolved_holder_name = await resolve_log_holder_name(
                log,
                operator.name if operator else None,
                db,
                holder_name_cache,
                user_name_cache
            )
            if holder_filter_keyword and holder_filter_keyword not in resolved_holder_name.lower():
                continue
            resolved_rows.append((log, equipment, model, operator, resolved_holder_name))

        # 写入数据
        for row_idx, (log, equipment, model, operator, resolved_holder_name) in enumerate(resolved_rows, 2):

            # 构建器材显示名称
            if equipment and model:
                equipment_display = f"{model.name} - {equipment.serial_number or '无编号'}"
            elif equipment:
                equipment_display = equipment.serial_number or "未知设备"
            elif log.action_type == "DELETE":
                equipment_display = "已删除设备"
            else:
                equipment_display = "未知设备"
            
            # 获取出库日期
            checkout_date = await get_checkout_date(log, db)
            
            data = [
                format_dt_in_timezone(log.created_at, user_tz, '%Y-%m-%d %H:%M:%S'),
                operator.name if operator else "未知用户",
                resolved_holder_name,
                map_action_type(log.action_type),
                equipment_display,
                format_dt_in_timezone(checkout_date, user_tz, '%Y-%m-%d %H:%M:%S'),
                log.purpose or "",
                format_dt_in_timezone(log.expected_return_at, user_tz, '%Y-%m-%d %H:%M:%S'),
                format_dt_in_timezone(log.actual_return_at, user_tz, '%Y-%m-%d %H:%M:%S'),
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
        
        # 调整列宽
        column_widths = [20, 15, 15, 12, 35, 20, 25, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        date_str = format_dt_in_timezone(datetime.utcnow(), user_tz, '%Y%m%d')
        filename = f"操作日志_{date_str}.xlsx"
        
        # 使用 urllib.parse.quote 处理中文文件名
        from urllib.parse import quote
        encoded_filename = quote(filename)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出日志失败: {str(e)}")


# ============ 操作日志 API ============

@app.get("/api/logs")
async def get_logs(
    request: Request,
    action_type: Optional[str] = None,
    equipment_id: Optional[int] = None,
    feishu_open_id: Optional[str] = None,
    serial_number: Optional[str] = None,
    model_name: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    operator_name: Optional[str] = None,
    holder_name: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    获取操作日志列表
    支持分页和多条件筛选
    新增：支持按设备编号(serial_number)和型号名称(model_name)筛选
    新增：支持出入库组合筛选
    """
    try:
        user_tz = get_request_timezone(request)
        start_date_utc = to_utc_naive(start_date, user_tz)
        end_date_utc = to_utc_naive(end_date, user_tz)

        # 操作类型映射函数 - 显示用
        def map_action_type(action_type: str) -> str:
            mapping = {
                "CHECKOUT": "出库",
                "ASSIGN": "分配",
                "CHECKIN": "入库",
                "TRANSFER": "交接",
                "CREATE": "添加设备",
                "UPDATE": "修改设备",
                "DELETE": "删除设备",
                "CREATE_CATEGORY": "添加分类",
                "UPDATE_CATEGORY": "修改分类",
                "DELETE_CATEGORY": "删除分类",
                "CREATE_MODEL": "添加型号",
                "UPDATE_MODEL": "修改型号",
                "DELETE_MODEL": "删除型号",
            }
            return mapping.get(action_type, action_type)
        
        # 所有编辑类操作类型（用于"编辑信息"筛选）
        edit_action_types = [
            "CREATE", "UPDATE", "DELETE",  # 设备操作
            "CREATE_CATEGORY", "UPDATE_CATEGORY", "DELETE_CATEGORY",  # 分类操作
            "CREATE_MODEL", "UPDATE_MODEL", "DELETE_MODEL"  # 型号操作
        ]
        # 出入库类型
        checkout_action_types = ["CHECKOUT", "ASSIGN", "CHECKIN", "TRANSFER"]
        
        # 构建查询 - 使用 outerjoin 以处理器材可能已被删除的情况
        # 同时关联 EquipmentModel 以获取型号信息
        query = select(OperationLog, Equipment, EquipmentModel, User).outerjoin(
            Equipment, OperationLog.equipment_id == Equipment.id
        ).outerjoin(
            EquipmentModel, Equipment.model_id == EquipmentModel.id
        ).join(
            User, OperationLog.feishu_open_id == User.feishu_open_id
        )
        
        # 应用筛选条件
        if action_type:
            if action_type == "EDIT":
                # 编辑信息 - 匹配设备增删改操作
                query = query.where(OperationLog.action_type.in_(edit_action_types))
            elif action_type == "出入库":
                # 出入库 - 同时匹配出库和入库
                query = query.where(OperationLog.action_type.in_(checkout_action_types))
            elif action_type in ["交接", "TRANSFER"]:
                query = query.where(OperationLog.action_type == "TRANSFER")
            elif action_type in ["分配", "ASSIGN"]:
                query = query.where(OperationLog.action_type == "ASSIGN")
            elif action_type == "出库":
                query = query.where(OperationLog.action_type == "CHECKOUT")
            elif action_type == "入库":
                query = query.where(OperationLog.action_type == "CHECKIN")
            else:
                # 单个操作类型筛选（包括具体的 CREATE_CATEGORY 等）
                query = query.where(OperationLog.action_type == action_type)
        if equipment_id:
            query = query.where(OperationLog.equipment_id == equipment_id)
        if feishu_open_id:
            query = query.where(OperationLog.feishu_open_id == feishu_open_id)
        if serial_number:
            query = query.where(Equipment.serial_number.ilike(f"%{serial_number}%"))
        if model_name:
            query = query.where(EquipmentModel.name.ilike(f"%{model_name}%"))
        if start_date_utc:
            query = query.where(OperationLog.created_at >= start_date_utc)
        if end_date_utc:
            query = query.where(OperationLog.created_at <= end_date_utc)
        if operator_name:
            query = query.where(User.name.ilike(f"%{operator_name}%"))
        if search:
            # 搜索关键词匹配设备型号名称、设备编号或用途
            search_pattern = f"%{search}%"
            query = query.where(
                or_(
                    EquipmentModel.name.ilike(search_pattern),
                    Equipment.serial_number.ilike(search_pattern),
                    OperationLog.purpose.ilike(search_pattern)
                )
            )
        
        holder_filter_keyword = holder_name.strip().lower() if holder_name else None

        if holder_filter_keyword:
            # 设备持有人由日志语义计算得出，无法直接用 SQL 准确筛选，先取全量后在内存中过滤
            full_result = await db.execute(query.order_by(desc(OperationLog.id)))
            full_rows = full_result.all()
            holder_name_cache: dict[int, str] = {}
            user_name_cache: dict[str, Optional[str]] = {}
            filtered_rows: list[tuple[OperationLog, Optional[Equipment], Optional[EquipmentModel], Optional[User], str]] = []
            for log, equipment, model, operator in full_rows:
                if operator and operator.feishu_open_id:
                    user_name_cache[operator.feishu_open_id] = operator.name

                resolved_holder_name = await resolve_log_holder_name(
                    log,
                    operator.name if operator else None,
                    db,
                    holder_name_cache,
                    user_name_cache
                )
                if holder_filter_keyword in resolved_holder_name.lower():
                    filtered_rows.append((log, equipment, model, operator, resolved_holder_name))

            total = len(filtered_rows)
            offset = (page - 1) * page_size
            page_rows = filtered_rows[offset: offset + page_size]
        else:
            # 获取总数（无持有人筛选时走 SQL 计数）
            count_query = select(OperationLog).outerjoin(
                Equipment, OperationLog.equipment_id == Equipment.id
            ).outerjoin(
                EquipmentModel, Equipment.model_id == EquipmentModel.id
            ).join(
                User, OperationLog.feishu_open_id == User.feishu_open_id
            )

            if action_type:
                if action_type == "EDIT":
                    count_query = count_query.where(OperationLog.action_type.in_(edit_action_types))
                elif action_type == "出入库":
                    count_query = count_query.where(OperationLog.action_type.in_(checkout_action_types))
                elif action_type in ["交接", "TRANSFER"]:
                    count_query = count_query.where(OperationLog.action_type == "TRANSFER")
                elif action_type in ["分配", "ASSIGN"]:
                    count_query = count_query.where(OperationLog.action_type == "ASSIGN")
                elif action_type == "出库":
                    count_query = count_query.where(OperationLog.action_type == "CHECKOUT")
                elif action_type == "入库":
                    count_query = count_query.where(OperationLog.action_type == "CHECKIN")
                else:
                    count_query = count_query.where(OperationLog.action_type == action_type)
            if equipment_id:
                count_query = count_query.where(OperationLog.equipment_id == equipment_id)
            if feishu_open_id:
                count_query = count_query.where(OperationLog.feishu_open_id == feishu_open_id)
            if serial_number:
                count_query = count_query.where(Equipment.serial_number.ilike(f"%{serial_number}%"))
            if model_name:
                count_query = count_query.where(EquipmentModel.name.ilike(f"%{model_name}%"))
            if start_date_utc:
                count_query = count_query.where(OperationLog.created_at >= start_date_utc)
            if end_date_utc:
                count_query = count_query.where(OperationLog.created_at <= end_date_utc)
            if operator_name:
                count_query = count_query.where(User.name.ilike(f"%{operator_name}%"))
            if search:
                # 搜索关键词匹配设备型号名称、设备编号或用途
                search_pattern = f"%{search}%"
                count_query = count_query.where(
                    or_(
                        EquipmentModel.name.ilike(search_pattern),
                        Equipment.serial_number.ilike(search_pattern),
                        OperationLog.purpose.ilike(search_pattern)
                    )
                )

            count_result = await db.execute(count_query)
            total = len(count_result.scalars().all())

            # 分页排序
            result = await db.execute(
                query.order_by(desc(OperationLog.id)).offset((page - 1) * page_size).limit(page_size)
            )
            rows = result.all()

            holder_name_cache: dict[int, str] = {}
            user_name_cache: dict[str, Optional[str]] = {}
            page_rows = []
            for log, equipment, model, operator in rows:
                if operator and operator.feishu_open_id:
                    user_name_cache[operator.feishu_open_id] = operator.name

                resolved_holder_name = await resolve_log_holder_name(
                    log,
                    operator.name if operator else None,
                    db,
                    holder_name_cache,
                    user_name_cache
                )
                page_rows.append((log, equipment, model, operator, resolved_holder_name))
        
        # 组装响应数据
        data = []
        for log, equipment, model, operator, resolved_holder_name in page_rows:

            # 构建器材显示名称：型号名称 + 设备编号
            if equipment and model:
                equipment_display = f"{model.name} - {equipment.serial_number or '无编号'}"
            elif equipment:
                equipment_display = equipment.serial_number or "未知设备"
            elif log.action_type == "DELETE":
                equipment_display = "已删除设备"
            else:
                equipment_display = "未知设备"
            
            data.append({
                "id": log.id,
                "equipment_id": log.equipment_id,
                "equipment_serial": equipment.serial_number if equipment else None,
                "equipment_model_name": model.name if model else None,
                "equipment_display": equipment_display,
                "feishu_open_id": log.feishu_open_id,
                "feishu_user_id": operator.feishu_user_id if operator else None,
                "user_name": operator.name if operator else None,
                "holder_name": resolved_holder_name,
                "action_type": map_action_type(log.action_type),
                "raw_action_type": log.action_type,  # 原始操作类型，供内部使用
                "purpose": log.purpose,
                "expected_return_at": log.expected_return_at,
                "actual_return_at": log.actual_return_at,
                "created_at": log.created_at
            })
        
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "data": data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取日志列表失败: {str(e)}")


@app.get("/api/logs/equipment/{equipment_id}")
async def get_equipment_logs(
    equipment_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    获取指定器材的操作日志
    """
    try:
        # 检查器材是否存在
        result = await db.execute(
            select(Equipment).where(Equipment.id == equipment_id)
        )
        if not result.scalar_one_or_none():
            raise HTTPException(status_code=404, detail="器材不存在")
        
        # 查询日志
        query = select(OperationLog, User).join(
            User, OperationLog.feishu_open_id == User.feishu_open_id
        ).where(OperationLog.equipment_id == equipment_id).order_by(desc(OperationLog.id))
        
        # 获取总数
        count_result = await db.execute(
            select(OperationLog).where(OperationLog.equipment_id == equipment_id)
        )
        total = len(count_result.scalars().all())
        
        # 分页
        query = query.offset((page - 1) * page_size).limit(page_size)
        result = await db.execute(query)
        rows = result.all()
        
        data = []
        for log, operator in rows:
            data.append({
                "id": log.id,
                "action_type": log.action_type,
                "feishu_open_id": log.feishu_open_id,
                "user_name": operator.name if operator else None,
                "purpose": log.purpose,
                "expected_return_at": log.expected_return_at,
                "actual_return_at": log.actual_return_at,
                "created_at": log.created_at
            })
        
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "data": data
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取器材日志失败: {str(e)}")


@app.get("/api/users")
async def get_users(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    获取用户列表
    """
    try:
        # 获取总数
        count_result = await db.execute(select(User))
        total = len(count_result.scalars().all())
        
        # 分页查询
        query = select(User).order_by(User.id.desc()).offset((page - 1) * page_size).limit(page_size)
        result = await db.execute(query)
        users = result.scalars().all()
        
        data = [{
            "id": u.id,
            "name": u.name,
            "avatar_url": u.avatar_url,
            "feishu_open_id": u.feishu_open_id,
            "feishu_user_id": u.feishu_user_id,
            "created_at": u.created_at.isoformat() if u.created_at else None
        } for u in users]
        
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "data": data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取用户列表失败: {str(e)}")


@app.get("/api/logs/user/{feishu_open_id}")
async def get_user_logs(
    feishu_open_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(require_auth),
    db: AsyncSession = Depends(get_db)
):
    """
    获取指定用户的操作日志
    """
    try:
        # 检查用户是否存在
        result = await db.execute(
            select(User).where(User.feishu_open_id == feishu_open_id)
        )
        if not result.scalar_one_or_none():
            raise HTTPException(status_code=404, detail="用户不存在")
        
        # 查询日志
        query = select(OperationLog, Equipment).join(
            Equipment, OperationLog.equipment_id == Equipment.id
        ).where(OperationLog.feishu_open_id == feishu_open_id).order_by(desc(OperationLog.id))
        
        # 获取总数
        count_result = await db.execute(
            select(OperationLog).where(OperationLog.feishu_open_id == feishu_open_id)
        )
        total = len(count_result.scalars().all())
        
        # 分页
        query = query.offset((page - 1) * page_size).limit(page_size)
        result = await db.execute(query)
        rows = result.all()
        
        data = []
        for log, equipment in rows:
            data.append({
                "id": log.id,
                "equipment_id": log.equipment_id,
                "equipment_serial": equipment.serial_number if equipment else None,
                "action_type": log.action_type,
                "purpose": log.purpose,
                "expected_return_at": log.expected_return_at,
                "actual_return_at": log.actual_return_at,
                "created_at": log.created_at
            })
        
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "data": data
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取用户日志失败: {str(e)}")


# ============ 主入口 ============

if __name__ == "__main__":
    import uvicorn
    import sys
    
    port = int(os.getenv("PORT", 8001))
    if len(sys.argv) > 1:
        try:
            port = int(sys.argv[1])
        except ValueError:
            pass
    
    print(f"🚀 启动飞书器材管理系统后端服务...")
    print(f"📍 访问地址: http://localhost:{port}")
    print(f"📚 API 文档: http://localhost:{port}/docs")
    
    uvicorn.run("main:app", host="0.0.0.0", port=port)
